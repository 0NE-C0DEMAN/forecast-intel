from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from streamlit.components.v1 import html as components_html

ROOT = Path(__file__).parent
FRONTEND = ROOT / "frontend"
DEFAULT_DATA = ROOT / "Forecast_26_Jan_Feb_Results_v4.xlsx"
MONTHLY_SHEET = "Monthly_Predictions"
MAPE_SHEET = "MAPE_Summary"

REQUIRED_COLUMNS = [
    "Item Code",
    "Item Description",
    "Is HV",
    "Tier",
    "Period",
    "Prev Closing Balance",
    "Predicted Closing Bal",
    "Actual Closing Bal",
    "Diff of Actual to Prediction",
    "Difference",
    "Predicted Action",
    "Actual Action",
    "Direction Correct",
    "Quantity",
    "Bias Correction Applied",
    "APE (%)",
    "Item MAPE (%)",
    "Months in MAPE",
]

st.set_page_config(
    page_title="Forecast Intel",
    page_icon=":bar_chart:",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ---------------------------------------------------------------------------
# Password gate
# ---------------------------------------------------------------------------
def _expected_password() -> str | None:
    """Read the gate password from st.secrets. Supports either top-level
    `password = "..."` or `[auth] password = "..."`."""
    try:
        if "password" in st.secrets:
            return str(st.secrets["password"])
        if "auth" in st.secrets and "password" in st.secrets["auth"]:
            return str(st.secrets["auth"]["password"])
    except Exception:
        pass
    return None


_LOGIN_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width,initial-scale=1.0" />
<title>Forecast Intel · Sign in</title>
<link rel="preconnect" href="https://fonts.googleapis.com" />
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800&display=swap" rel="stylesheet" />
<script>
/* Match the app: bump non-Windows screens (Mac etc.) up to 125% so the UI
   isn't tiny there. Windows (usually 125% OS scaling) is left at 100%. */
(function () {
  try {
    var ua = navigator.userAgent || '', plat = navigator.platform || '';
    if (!(/Win/i.test(plat) || /Windows/i.test(ua))) document.documentElement.style.zoom = '1.25';
  } catch (e) {}
})();
</script>
<script src="https://unpkg.com/react@18/umd/react.production.min.js"></script>
<script src="https://unpkg.com/react-dom@18/umd/react-dom.production.min.js"></script>
<script src="https://unpkg.com/@babel/standalone/babel.min.js"></script>
<style>
  :root {
    --accent: #4F46E5;
    --accent-2: #6366F1;
    --accent-3: #818CF8;
    --text: #111827;
    --text-2: #6B7280;
    --text-3: #9CA3AF;
    --border: #E5E7EB;
    --surface: #FFFFFF;
    --surface-2: #F7F8FA;
    --danger: #DC2626;
    --danger-bg: rgba(220,38,38,.06);
    --danger-bd: rgba(220,38,38,.18);
  }
  html, body, #root { margin:0; padding:0; height:100%; min-height:100vh; }
  body {
    font-family: 'Outfit', -apple-system, system-ui, 'Segoe UI', sans-serif;
    color: var(--text);
    background:
      radial-gradient(900px 600px at 12% -10%, rgba(99,102,241,.18), transparent 60%),
      radial-gradient(700px 500px at 110% 110%, rgba(79,70,229,.14), transparent 60%),
      linear-gradient(180deg, #F7F8FA 0%, #EEF2FF 100%);
    overflow: hidden;
  }
  * { box-sizing: border-box; }
  @keyframes lg-spin { to { transform: rotate(360deg); } }
  @keyframes lg-shake {
    0%,100% { transform: translateX(0); }
    20% { transform: translateX(-7px); }
    40% { transform: translateX(7px); }
    60% { transform: translateX(-4px); }
    80% { transform: translateX(4px); }
  }
  @keyframes lg-rise {
    from { opacity: 0; transform: translateY(8px); }
    to { opacity: 1; transform: translateY(0); }
  }
  .lg-shake { animation: lg-shake .42s; }
  .lg-card { animation: lg-rise .38s ease-out both; }
  ::placeholder { color: #9CA3AF; }
</style>
</head>
<body>
<div id="root"></div>
<script>
  window.__submitLogin = function(pw) {
    try {
      var doc = window.parent.document;
      var input = doc.querySelector('[data-testid="stFileUploaderDropzoneInput"]')
        || doc.querySelector('[data-testid="stFileUploader"] input[type="file"]')
        || doc.querySelector('input[type="file"]');
      if (!input) return false;
      var blob = new Blob([pw], { type: 'text/csv' });
      var file = new File([blob], '__LOGIN_ATTEMPT__.csv', { type: 'text/csv' });
      var dt = new DataTransfer();
      dt.items.add(file);
      var setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'files').set;
      setter.call(input, dt.files);
      input.dispatchEvent(new Event('change', { bubbles: true }));
      return true;
    } catch (e) { console.error('login bridge failed', e); return false; }
  };
</script>
<script type="text/babel">
const { useState, useRef, useEffect } = React;

function LockIcon() {
  return (
    <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="#6B7280" strokeWidth="2" strokeLinecap="round" strokeLinejoin="round">
      <rect x="3" y="11" width="18" height="11" rx="2"></rect>
      <path d="M7 11V7a5 5 0 0 1 10 0v4"></path>
    </svg>
  );
}

function EyeIcon({ open }) {
  return open ? (
    <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" strokeWidth="2" strokeLinecap="round" strokeLinejoin="round">
      <path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8S1 12 1 12z"></path>
      <circle cx="12" cy="12" r="3"></circle>
    </svg>
  ) : (
    <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" strokeWidth="2" strokeLinecap="round" strokeLinejoin="round">
      <path d="M17.94 17.94A10.07 10.07 0 0 1 12 20c-7 0-11-8-11-8a18.45 18.45 0 0 1 5.06-5.94M9.9 4.24A9.12 9.12 0 0 1 12 4c7 0 11 8 11 8a18.5 18.5 0 0 1-2.16 3.19m-6.72-1.07a3 3 0 1 1-4.24-4.24"></path>
      <line x1="1" y1="1" x2="23" y2="23"></line>
    </svg>
  );
}

function AlertIcon() {
  return (
    <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#B91C1C" strokeWidth="2.4" strokeLinecap="round" strokeLinejoin="round">
      <circle cx="12" cy="12" r="10"></circle>
      <line x1="12" y1="8" x2="12" y2="13"></line>
      <line x1="12" y1="16" x2="12.01" y2="16"></line>
    </svg>
  );
}

function Spinner() {
  return (
    <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="#fff" strokeWidth="2.4" strokeLinecap="round" style={{ animation: 'lg-spin .7s linear infinite' }}>
      <path d="M21 12a9 9 0 1 1-6.22-8.56"></path>
    </svg>
  );
}

function LoginScreen({ initialError }) {
  const [pw, setPw] = useState('');
  const [show, setShow] = useState(false);
  const [busy, setBusy] = useState(false);
  const [err, setErr] = useState(initialError || null);
  const [shake, setShake] = useState(!!initialError);
  const [focused, setFocused] = useState(false);
  const inputRef = useRef(null);

  useEffect(() => {
    const t = setTimeout(() => inputRef.current && inputRef.current.focus(), 80);
    if (initialError) {
      const t2 = setTimeout(() => setShake(false), 450);
      return () => { clearTimeout(t); clearTimeout(t2); };
    }
    return () => clearTimeout(t);
  }, []);

  function submit(e) {
    e.preventDefault();
    if (!pw || busy) return;
    setBusy(true);
    setErr(null);
    const ok = window.__submitLogin(pw);
    if (!ok) {
      setBusy(false);
      setErr('Bridge not ready. Refresh the page and try again.');
      setShake(true);
      setTimeout(() => setShake(false), 450);
    }
    // on success, the page reruns and remounts authed; busy stays true visually until then
  }

  const disabled = busy || !pw;
  const inputBorder = err ? '#FCA5A5' : (focused ? 'var(--accent)' : 'var(--border)');
  const inputBg = focused ? '#FFFFFF' : 'var(--surface-2)';
  const inputShadow = focused ? '0 0 0 3px rgba(79,70,229,0.12)' : 'none';

  return (
    <div style={{
      minHeight: '100vh', width: '100%',
      display: 'flex', alignItems: 'center', justifyContent: 'center',
      padding: 20,
    }}>
      <form
        onSubmit={submit}
        className={'lg-card' + (shake ? ' lg-shake' : '')}
        style={{
          width: '100%', maxWidth: 420,
          background: 'var(--surface)',
          border: '1px solid var(--border)',
          borderRadius: 20,
          padding: '34px 36px 28px',
          boxShadow: '0 28px 64px rgba(15,23,42,0.10), 0 8px 16px rgba(15,23,42,0.05)',
          position: 'relative', zIndex: 1,
        }}
      >
        {/* Brand row */}
        <div style={{ display: 'flex', alignItems: 'center', gap: 12, marginBottom: 26 }}>
          <div style={{
            width: 44, height: 44, borderRadius: 12,
            background: 'linear-gradient(135deg, var(--accent), var(--accent-3))',
            display: 'flex', alignItems: 'center', justifyContent: 'center',
            color: '#fff', fontWeight: 800, fontSize: 19, letterSpacing: '-0.02em',
            boxShadow: '0 8px 18px rgba(79,70,229,0.32)'
          }}>P</div>
          <div style={{ lineHeight: 1.15 }}>
            <div style={{ fontSize: 16, fontWeight: 800, color: 'var(--text)', letterSpacing: '-0.01em' }}>Forecast Intel</div>
            <div style={{ fontSize: 12, color: 'var(--text-2)', marginTop: 2 }}>Monthly inventory predictions</div>
          </div>
        </div>

        {/* Heading */}
        <div style={{ fontSize: 22, fontWeight: 800, color: 'var(--text)', letterSpacing: '-0.02em', marginBottom: 6 }}>Welcome back</div>
        <div style={{ fontSize: 13, color: 'var(--text-2)', lineHeight: 1.55, marginBottom: 22 }}>
          Sign in with the access password to view forecasts, accuracy metrics, and SKU-level predictions.
        </div>

        {/* Password label */}
        <label htmlFor="lg-pw" style={{
          display: 'block', fontSize: 11, fontWeight: 600, color: '#4B5563',
          textTransform: 'uppercase', letterSpacing: '0.07em', marginBottom: 7
        }}>Access password</label>

        {/* Password input */}
        <div style={{ position: 'relative' }}>
          <span style={{ position: 'absolute', left: 12, top: '50%', transform: 'translateY(-50%)', display: 'flex', pointerEvents: 'none' }}>
            <LockIcon />
          </span>
          <input
            id="lg-pw"
            ref={inputRef}
            type={show ? 'text' : 'password'}
            value={pw}
            onChange={e => setPw(e.target.value)}
            placeholder="••••••••"
            disabled={busy}
            autoComplete="current-password"
            style={{
              width: '100%',
              padding: '11px 42px 11px 38px',
              background: inputBg,
              border: '1px solid ' + inputBorder,
              borderRadius: 10,
              fontSize: 14,
              fontFamily: "'Outfit', system-ui, sans-serif",
              color: 'var(--text)',
              outline: 'none',
              transition: 'border-color .15s, background .15s, box-shadow .15s',
              boxShadow: inputShadow,
            }}
            onFocus={() => setFocused(true)}
            onBlur={() => setFocused(false)}
          />
          <button
            type="button"
            onClick={() => setShow(s => !s)}
            tabIndex={-1}
            aria-label={show ? 'Hide password' : 'Show password'}
            style={{
              position: 'absolute', right: 8, top: '50%', transform: 'translateY(-50%)',
              background: 'transparent', border: 'none', cursor: 'pointer',
              padding: 6, borderRadius: 6, display: 'flex',
              color: focused ? 'var(--accent)' : '#9CA3AF',
              transition: 'color .15s, background .15s'
            }}
            onMouseEnter={e => e.currentTarget.style.background = '#F3F4F6'}
            onMouseLeave={e => e.currentTarget.style.background = 'transparent'}
          >
            <EyeIcon open={show} />
          </button>
        </div>

        {/* Error banner */}
        {err && (
          <div style={{
            marginTop: 12, padding: '10px 12px', borderRadius: 10,
            background: 'var(--danger-bg)', border: '1px solid var(--danger-bd)',
            color: '#B91C1C', fontSize: 12.5, fontWeight: 600,
            display: 'flex', alignItems: 'center', gap: 8, lineHeight: 1.45
          }}>
            <AlertIcon />
            <span>{err}</span>
          </div>
        )}

        {/* Submit button */}
        <button
          type="submit"
          disabled={disabled}
          style={{
            marginTop: 18, width: '100%',
            padding: '12px 16px', borderRadius: 10,
            background: disabled
              ? '#E5E7EB'
              : 'linear-gradient(135deg, var(--accent), var(--accent-2))',
            color: disabled ? '#9CA3AF' : '#fff', border: 'none',
            cursor: disabled ? 'not-allowed' : 'pointer',
            fontWeight: 700, fontSize: 14,
            fontFamily: "'Outfit', system-ui, sans-serif",
            letterSpacing: '0.01em',
            boxShadow: disabled ? 'none' : '0 8px 18px rgba(79,70,229,0.30)',
            transition: 'transform .12s ease, box-shadow .12s ease, background .12s',
            display: 'flex', alignItems: 'center', justifyContent: 'center', gap: 9,
          }}
          onMouseEnter={e => {
            if (disabled) return;
            e.currentTarget.style.transform = 'translateY(-1px)';
            e.currentTarget.style.boxShadow = '0 12px 24px rgba(79,70,229,0.38)';
          }}
          onMouseLeave={e => {
            if (disabled) return;
            e.currentTarget.style.transform = 'translateY(0)';
            e.currentTarget.style.boxShadow = '0 8px 18px rgba(79,70,229,0.30)';
          }}
        >
          {busy ? (<><Spinner />Verifying…</>) : 'Sign in'}
        </button>

        {/* Footer */}
        <div style={{
          marginTop: 22, paddingTop: 18,
          borderTop: '1px solid #F3F4F6',
          textAlign: 'center', fontSize: 11.5, color: 'var(--text-3)'
        }}>
          Contact the project owner if you don't have a password.
        </div>
      </form>
    </div>
  );
}

const initialError = window.__INITIAL_ERROR__ || null;
ReactDOM.createRoot(document.getElementById('root')).render(<LoginScreen initialError={initialError} />);
</script>
</body>
</html>
"""


def _render_login_screen(error: str | None = None) -> None:
    # Hide all Streamlit chrome behind the iframe and let it own the full viewport.
    st.markdown(
        """
<style>
[data-testid="stHeader"], [data-testid="stToolbar"], #MainMenu, footer,
[data-testid="stStatusWidget"], [data-testid="stDeployButton"] { display: none !important; height: 0 !important; }
section[data-testid="stSidebar"] { display: none !important; }
.stApp > header { display: none !important; }
.stApp { background: #EEF2FF; }
.block-container, [data-testid="stAppViewBlockContainer"] { padding: 0 !important; max-width: 100% !important; }
[data-testid="stMain"], [data-testid="stAppViewContainer"], [data-testid="stMainBlockContainer"] {
  height: 100vh !important; max-height: 100vh !important; overflow: hidden !important; padding: 0 !important;
}
[data-testid="stMainBlockContainer"], [data-testid="stVerticalBlock"], [data-testid="stElementContainer"] {
  padding: 0 !important; margin: 0 !important; gap: 0 !important;
}
iframe[data-testid="stIFrame"] {
  width: 100vw !important; height: 100vh !important; border: 0 !important;
  margin: 0 !important; padding: 0 !important; display: block !important;
}
[data-testid="stElementContainer"]:has(> iframe[data-testid="stIFrame"]) {
  height: 100vh !important; min-height: 0 !important; max-height: 100vh !important;
}
html, body { overflow: hidden !important; height: 100vh !important; }
.st-key-forecast_bridge {
  position: fixed !important; left: -10000px !important; top: 0 !important;
  width: 1px !important; height: 1px !important; overflow: hidden !important;
  pointer-events: none !important; opacity: 0 !important;
}
</style>
""",
        unsafe_allow_html=True,
    )

    # Inject initial error (if any) so the JSX can show it on first render
    err_js = json.dumps(error) if error else "null"
    html_doc = _LOGIN_HTML.replace(
        "<script type=\"text/babel\">",
        f"<script>window.__INITIAL_ERROR__ = {err_js};</script>\n<script type=\"text/babel\">",
        1,
    )
    components_html(html_doc, height=900, scrolling=False)


_DASHBOARD_CSS = """
<style>
[data-testid="stHeader"], [data-testid="stToolbar"], #MainMenu, footer,
[data-testid="stStatusWidget"], [data-testid="stDeployButton"] { display: none !important; height: 0 !important; }
section[data-testid="stSidebar"] { display: none !important; }
.stApp { background: #FAFBFC; }
.block-container, [data-testid="stAppViewBlockContainer"] { padding: 0 !important; max-width: 100% !important; }
.stApp > header { display: none !important; }
iframe[data-testid="stIFrame"] {
  width: 100vw !important;
  height: 100vh !important;
  border: 0 !important;
  margin: 0 !important;
  padding: 0 !important;
  display: block !important;
}
[data-testid="stElementContainer"]:has(> iframe[data-testid="stIFrame"]) {
  height: 100vh !important;
  min-height: 0 !important;
  max-height: 100vh !important;
}
[data-testid="stMainBlockContainer"], [data-testid="stVerticalBlock"], [data-testid="stElementContainer"] {
  padding: 0 !important;
  margin: 0 !important;
  gap: 0 !important;
}
[data-testid="stMain"], [data-testid="stAppViewContainer"], [data-testid="stMainBlockContainer"] {
  height: 100vh !important;
  max-height: 100vh !important;
  overflow: hidden !important;
}
html, body { overflow: hidden !important; height: 100vh !important; }
.st-key-forecast_bridge {
  position: fixed !important;
  left: -10000px !important; top: 0 !important;
  width: 1px !important; height: 1px !important;
  overflow: hidden !important; pointer-events: none !important;
  opacity: 0 !important;
}
</style>
"""


def _df_to_records(df: pd.DataFrame) -> list[dict]:
    def num(v):
        if pd.isna(v):
            return None
        return float(v)

    def text(v):
        if pd.isna(v):
            return None
        return str(v).strip()

    def boolish(v):
        if pd.isna(v):
            return None
        return bool(v)

    def intish(v):
        if pd.isna(v):
            return None
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    records: list[dict] = []
    for _, r in df.iterrows():
        records.append(
            {
                "itemCode": text(r.get("Item Code")),
                "description": text(r.get("Item Description")),
                "isHV": bool(r.get("Is HV")) if pd.notna(r.get("Is HV")) else False,
                "tier": text(r.get("Tier")),
                "period": text(r.get("Period")),
                "prevClosingBal": num(r.get("Prev Closing Balance")),
                "predictedClosingBal": num(r.get("Predicted Closing Bal")),
                "actualClosingBal": num(r.get("Actual Closing Bal")),
                # 'error' is the per-row Actual − Predicted delta (renamed from "Error" to "Diff of Actual to Prediction" in v4)
                "error": num(r.get("Diff of Actual to Prediction", r.get("Error"))),
                "difference": num(r.get("Difference")),
                "predictedAction": text(r.get("Predicted Action")),
                "actualAction": text(r.get("Actual Action")),
                "directionCorrect": boolish(r.get("Direction Correct")),
                "quantity": num(r.get("Quantity")),
                "biasCorrection": num(r.get("Bias Correction Applied")),
                "ape": num(r.get("APE (%)")),
                "itemMape": num(r.get("Item MAPE (%)")),
                "mapeMonths": intish(r.get("Months in MAPE")),
            }
        )
    return records


def _mape_df_to_records(df: pd.DataFrame) -> list[dict]:
    def num(v):
        if pd.isna(v):
            return None
        return float(v)

    def text(v):
        if pd.isna(v):
            return None
        return str(v).strip()

    def intval(v):
        if pd.isna(v):
            return None
        return int(v)

    records = []
    for _, r in df.iterrows():
        records.append(
            {
                "period": text(r.get("Period")),
                "model": text(r.get("Model")),
                "mapeAll": num(r.get("MAPE All Items (%)")),
                "mapeHV": num(r.get("MAPE HV Items (%)")),
                "itemsPredicted": intval(r.get("Items Predicted")),
                "itemsDeliver": intval(r.get("Items Deliver")),
                "itemsReturn": intval(r.get("Items Return")),
                "tier": text(r.get("Tier")),
            }
        )
    return records


@st.cache_data(show_spinner=False)
def load_default_records(mtime: float) -> dict:
    """Load all sheets from the default workbook. Returns dict keyed by model type."""
    sheets = pd.read_excel(DEFAULT_DATA, sheet_name=None)

    def safe_load(sheet_name: str) -> list[dict]:
        if sheet_name in sheets:
            return _df_to_records(sheets[sheet_name])
        return []

    monthly = safe_load(MONTHLY_SHEET)

    mape_records: list[dict] = []
    if MAPE_SHEET in sheets:
        all_mape = _mape_df_to_records(sheets[MAPE_SHEET])
        # Only keep monthly rows — quarterly/half-yearly not used
        mape_records = [r for r in all_mape if r.get("model", "").lower().startswith("month")]

    return {
        "monthly": monthly,
        "mape": mape_records,
    }


@st.cache_data(show_spinner=False)
def parse_uploaded_records(blob: bytes, name: str) -> tuple[list[dict] | None, list[str], list[dict]]:
    issues: list[str] = []
    mape_records: list[dict] = []
    try:
        if name.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(blob))
        else:
            file_sheets = pd.read_excel(io.BytesIO(blob), sheet_name=None)
            # Try to extract MAPE summary if present
            if MAPE_SHEET in file_sheets:
                try:
                    all_mape = _mape_df_to_records(file_sheets[MAPE_SHEET])
                    mape_records = [r for r in all_mape if r.get("model", "").lower().startswith("month")]
                except Exception:
                    pass
            if MONTHLY_SHEET in file_sheets:
                df = file_sheets[MONTHLY_SHEET]
            else:
                month_cands = [k for k in file_sheets if "month" in k.lower()]
                if not month_cands:
                    return None, [
                        f"No '{MONTHLY_SHEET}' sheet (or any sheet containing 'month') was found."
                    ], []
                df = file_sheets[month_cands[0]]
                issues.append(
                    f"Sheet '{MONTHLY_SHEET}' not found; using '{month_cands[0]}' instead."
                )
    except Exception as exc:
        return None, [f"Could not parse the file: {exc}"], []

    # Accept legacy alias: older files used "Error" instead of "Diff of Actual to Prediction".
    if "Diff of Actual to Prediction" not in df.columns and "Error" in df.columns:
        df = df.rename(columns={"Error": "Diff of Actual to Prediction"})
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    # 'Bias Correction Applied', 'APE (%)', 'Months in MAPE' are v4-only; skip them in legacy files.
    optional = {"Bias Correction Applied", "APE (%)", "Months in MAPE"}
    hard_missing = [c for c in missing if c not in optional]
    if hard_missing:
        return None, [f"Missing required columns: {', '.join(hard_missing)}"], []

    return _df_to_records(df), issues, mape_records


# ---------------------------------------------------------------------------
# Supabase data source — preferred over the bundled Excel when creds exist.
# Schema is documented in `Tecscon_Developer_Integration_Guide.docx`. Two
# tables: forecast_runs (one row per model run) and forecast_predictions
# (one row per item × period × run). The Supabase schema is leaner than the
# v4 Excel, so several fields are derived here: Tier from is_high_value,
# actualAction + directionCorrect from actual vs previous balance, Item MAPE
# and Months in MAPE aggregated across each item's per-period APE.
# ---------------------------------------------------------------------------
# Threshold below which a balance delta is treated as "No Change" rather
# than a Deliver/Return. Calibrated to match the model's labelling.
_ACTION_THRESHOLD = 0.5


def _supabase_creds() -> tuple[str | None, str | None]:
    """Read SUPABASE_URL and SUPABASE_KEY from st.secrets. Returns (url, key)
    or (None, None) if either is missing — caller should fall back to Excel."""
    try:
        url = st.secrets.get("supabase_url") or st.secrets.get("SUPABASE_URL")
        key = st.secrets.get("supabase_key") or st.secrets.get("SUPABASE_KEY")
    except Exception:
        return None, None
    return (str(url) if url else None, str(key) if key else None)


@st.cache_resource(show_spinner=False)
def _supabase_client(url: str, key: str):
    """Create-once Supabase client. Cached across reruns via st.cache_resource."""
    from supabase import create_client  # local import — avoids hard dep if not used
    return create_client(url, key)


def _classify_action(delta: float | None) -> str | None:
    """Map a (current − previous) balance delta onto Deliver/Return/No Change
    using the same threshold the upstream model appears to use."""
    if delta is None:
        return None
    if delta >= _ACTION_THRESHOLD:
        return "Deliver"
    if delta <= -_ACTION_THRESHOLD:
        return "Return"
    return "No Change"


@st.cache_data(show_spinner=False, ttl=300)
def _fetch_supabase_runs(url: str, key: str) -> list[dict]:
    """Return all runs ordered newest → oldest. Cached briefly so the run
    selector doesn't hammer Supabase on every interaction."""
    client = _supabase_client(url, key)
    resp = (
        client.table("forecast_runs")
        .select("*")
        .order("run_timestamp", desc=True)
        .execute()
    )
    return list(resp.data or [])


@st.cache_data(show_spinner=False, ttl=300)
def _fetch_supabase_predictions(url: str, key: str, run_id: int) -> list[dict]:
    """Paginated fetch of every prediction row for a run (Supabase caps each
    request at 1000 rows; runs hold 2-3k)."""
    client = _supabase_client(url, key)
    all_rows: list[dict] = []
    offset = 0
    PAGE = 1000
    while True:
        resp = (
            client.table("forecast_predictions")
            .select("*")
            .eq("run_id", run_id)
            .order("year_month")
            .range(offset, offset + PAGE - 1)
            .execute()
        )
        page = list(resp.data or [])
        if not page:
            break
        all_rows.extend(page)
        if len(page) < PAGE:
            break
        offset += PAGE
    return all_rows


def _supabase_to_records(pred_rows: list[dict]) -> list[dict]:
    """Convert Supabase prediction rows into the JSON record shape the React
    UI expects.

    As of 2026-05-28 the schema stores every Excel column directly, so we no
    longer derive Tier / Difference / Actual Action / Direction Correct /
    Item MAPE / Months in MAPE / Bias Correction on the client. We just map
    the column names through. The per-row `forecast_mode` flag is also
    carried over so the UI can render a continuous Backtest+Future view."""
    if not pred_rows:
        return []

    def fnum(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def ftext(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        return str(v).strip()

    def fbool(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        return bool(v)

    def fint(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    records: list[dict] = []
    for r in pred_rows:
        records.append({
            "itemCode": ftext(r.get("item_code")),
            "description": ftext(r.get("item_description")),
            "isHV": bool(r.get("is_high_value")),
            "tier": ftext(r.get("tier")),
            "period": ftext(r.get("year_month")),
            "prevClosingBal": fnum(r.get("prev_closing_bal")),
            "predictedClosingBal": fnum(r.get("pred_closing_balance")),
            "actualClosingBal": fnum(r.get("actual_closing_balance")),
            "error": fnum(r.get("diff_actual_to_prediction")),
            "difference": fnum(r.get("difference")),
            "predictedAction": ftext(r.get("pred_action")),
            "actualAction": ftext(r.get("actual_action")),
            "directionCorrect": fbool(r.get("direction_correct")),
            "quantity": fnum(r.get("action_quantity")),
            "biasCorrection": fnum(r.get("bias_correction_applied")),
            "ape": fnum(r.get("ape")),
            "itemMape": fnum(r.get("item_mape")),
            "mapeMonths": fint(r.get("months_in_mape")),
            "forecastMode": ftext(r.get("forecast_mode")),
        })
    return records


def _synth_mape_summary(pred_rows: list[dict], run: dict | None) -> list[dict]:
    """Build the per-period MAPE summary the Model Accuracy page expects.
    Supabase doesn't have a MAPE_Summary table — we compute it from the
    predictions: one row per period with MAPE All / MAPE HV / item counts.

    Per-row APE column is `ape` in the 2026-05-28 schema (was `mape` in
    the older schema; we accept either to stay defensive)."""
    if not pred_rows:
        return []
    df = pd.DataFrame(pred_rows)
    if "year_month" not in df.columns:
        return []
    # Locate the per-row error column. New schema = `ape`; old = `mape`.
    err_col = "ape" if "ape" in df.columns else ("mape" if "mape" in df.columns else None)
    tier_label = (run or {}).get("forecast_mode") or "Monthly"
    out: list[dict] = []
    for period, sub in df.groupby("year_month"):
        if err_col is not None:
            all_mape = sub[err_col].dropna()
            hv_mape = sub.loc[sub["is_high_value"] == 1, err_col].dropna() if "is_high_value" in sub.columns else pd.Series(dtype=float)
        else:
            all_mape = pd.Series(dtype=float)
            hv_mape = pd.Series(dtype=float)
        deliver = int((sub.get("pred_action") == "Deliver").sum()) if "pred_action" in sub.columns else 0
        ret = int((sub.get("pred_action") == "Return").sum()) if "pred_action" in sub.columns else 0
        out.append({
            "period": str(period),
            "model": "Monthly",
            "mapeAll": float(all_mape.mean()) if len(all_mape) else None,
            "mapeHV": float(hv_mape.mean()) if len(hv_mape) else None,
            "itemsPredicted": int(len(sub)),
            "itemsDeliver": deliver,
            "itemsReturn": ret,
            "tier": tier_label,
        })
    out.sort(key=lambda r: r["period"])
    return out


@st.cache_data(show_spinner=False, ttl=300)
def _fetch_all_supabase_predictions(url: str, key: str) -> list[dict]:
    """Paginated fetch of every prediction row across every run. The
    dashboard now renders a single continuous view (no per-year split);
    Backtest and Future rows are distinguished by the per-row
    `forecast_mode` column rather than by which year/run they belong to."""
    client = _supabase_client(url, key)
    all_rows: list[dict] = []
    offset = 0
    PAGE = 1000
    while True:
        resp = (
            client.table("forecast_predictions")
            .select("*")
            .order("year_month")
            .range(offset, offset + PAGE - 1)
            .execute()
        )
        page = list(resp.data or [])
        if not page:
            break
        all_rows.extend(page)
        if len(page) < PAGE:
            break
        offset += PAGE
    return all_rows


def _load_supabase_all(url: str, key: str, runs: list[dict] | None = None) -> tuple[list[dict], list[dict], dict | None]:
    """Fetch every prediction row across every run and return
    (records, mape_summary, meta). The meta object is a synthetic
    "run" used only for the source label + MAPE-summary tier text."""
    pred_rows = _fetch_all_supabase_predictions(url, key)
    records = _supabase_to_records(pred_rows)
    if runs is None:
        runs = _fetch_supabase_runs(url, key)
    periods = sorted({str(p.get("year_month")) for p in pred_rows if p.get("year_month")})
    modes = sorted({str(p.get("forecast_mode")) for p in pred_rows if p.get("forecast_mode")})
    synth_meta = {
        "id": "all",
        "forecast_mode": "/".join(modes) if modes else None,
        "predict_year": None,
        "run_timestamp": "",
        "period_count": len(periods),
        "row_count": len(pred_rows),
        "periods": periods,
        "modes": modes,
        "runs": runs or [],
    }
    mape_summary = _synth_mape_summary(pred_rows, synth_meta)
    return records, mape_summary, synth_meta


def _supabase_all_source_label(meta: dict | None) -> str:
    if not meta:
        return "Supabase"
    bits = ["All forecasts"]
    pc = meta.get("period_count")
    if pc:
        bits.append(f"{pc} mo")
    modes = meta.get("modes") or []
    if modes:
        bits.append("/".join(sorted(set(modes))))
    return "Supabase · " + " · ".join(bits)


@st.cache_data(show_spinner=False, ttl=300)
def _fetch_supabase_predictions_by_year(url: str, key: str, year: int) -> list[dict]:
    """Paginated fetch of every prediction row whose `year_month` falls in
    the given year. Backtest and Future runs both contribute — the per-row
    `forecast_mode` column distinguishes them."""
    client = _supabase_client(url, key)
    all_rows: list[dict] = []
    offset = 0
    PAGE = 1000
    lo = f"{year}-01"
    hi = f"{year}-12"
    while True:
        resp = (
            client.table("forecast_predictions")
            .select("*")
            .gte("year_month", lo)
            .lte("year_month", hi)
            .order("year_month")
            .range(offset, offset + PAGE - 1)
            .execute()
        )
        page = list(resp.data or [])
        if not page:
            break
        all_rows.extend(page)
        if len(page) < PAGE:
            break
        offset += PAGE
    return all_rows


def _years_summary(runs: list[dict], predictions_by_year: dict[int, list[dict]] | None = None) -> list[dict]:
    """Group runs by `predict_year` and roll up each year's metadata for the
    sidebar picker. Optionally consumes already-fetched predictions per year
    to fill in exact period count + per-mode item counts; falls back to the
    run-level summary fields otherwise."""
    by_year: dict[int, dict] = {}
    for r in runs:
        y = r.get("predict_year")
        if y is None:
            continue
        y = int(y)
        slot = by_year.setdefault(y, {
            "year": y,
            "run_ids": [],
            "modes": [],
            "total_items": 0,
            "hv_item_count": 0,
            "avg_mape": None,
            "latest_ts": "",
            "training_years": "",
        })
        slot["run_ids"].append(int(r["id"]))
        if r.get("forecast_mode"):
            slot["modes"].append(r["forecast_mode"])
        # Pick the biggest item count among the year's runs as the headline
        ti = int(r.get("total_items") or 0)
        if ti > slot["total_items"]:
            slot["total_items"] = ti
            slot["hv_item_count"] = int(r.get("hv_item_count") or 0)
            slot["training_years"] = r.get("training_years") or ""
        # Latest timestamp
        ts = str(r.get("run_timestamp") or "")
        if ts > slot["latest_ts"]:
            slot["latest_ts"] = ts
        # Average MAPE: take backtest run's avg if present (Future runs have NULL)
        if r.get("avg_mape") is not None:
            slot["avg_mape"] = float(r["avg_mape"])

    if predictions_by_year:
        for y, preds in predictions_by_year.items():
            if y not in by_year:
                continue
            periods = sorted({str(p.get("year_month")) for p in preds if p.get("year_month")})
            by_year[y]["period_count"] = len(periods)
            by_year[y]["periods"] = periods
            by_year[y]["row_count"] = len(preds)

    return sorted(by_year.values(), key=lambda x: x["year"], reverse=True)


def _load_supabase_year(url: str, key: str, year: int, runs: list[dict] | None = None) -> tuple[list[dict], list[dict], dict | None]:
    """Fetch every prediction row for a given year (across all runs) and
    return (records, mape_summary, year_meta) for the dashboard to render."""
    pred_rows = _fetch_supabase_predictions_by_year(url, key, year)
    records = _supabase_to_records(pred_rows)
    if runs is None:
        runs = _fetch_supabase_runs(url, key)
    year_summary = next((y for y in _years_summary(runs, {year: pred_rows}) if y["year"] == int(year)), None)
    # Build a synthetic "run meta" object so the MAPE summary tier label is
    # informative (carries the dominant mode + the actual year).
    synth_run = None
    if year_summary:
        synth_run = {
            "id": ",".join(map(str, year_summary["run_ids"])),
            "forecast_mode": "/".join(sorted(set(year_summary["modes"]))) or None,
            "predict_year": year,
            "run_timestamp": year_summary["latest_ts"],
        }
    mape_summary = _synth_mape_summary(pred_rows, synth_run)
    return records, mape_summary, year_summary


def _load_supabase_run(url: str, key: str, run_id: int) -> tuple[list[dict], list[dict], dict | None]:
    """Single-run fetch (kept for back-compat / debugging; the year-based
    view is the primary UX now)."""
    runs = _fetch_supabase_runs(url, key)
    run_meta = next((r for r in runs if int(r.get("id", -1)) == int(run_id)), None)
    pred_rows = _fetch_supabase_predictions(url, key, run_id)
    records = _supabase_to_records(pred_rows)
    mape_summary = _synth_mape_summary(pred_rows, run_meta)
    return records, mape_summary, run_meta


def _supabase_year_source_label(year_meta: dict | None) -> str:
    if not year_meta:
        return "Supabase"
    bits = [f"Forecast {year_meta['year']}"]
    pc = year_meta.get("period_count")
    if pc:
        bits.append(f"{pc} mo")
    modes = year_meta.get("modes") or []
    if modes:
        bits.append("/".join(sorted(set(modes))))
    return "Supabase · " + " · ".join(bits)


def _supabase_source_label(run_meta: dict | None) -> str:
    if not run_meta:
        return "Supabase"
    ts = str(run_meta.get("run_timestamp") or "")[:16].replace("T", " ")
    mode = run_meta.get("forecast_mode") or ""
    pyear = run_meta.get("predict_year")
    bits = [f"Run #{run_meta.get('id')}"]
    if pyear:
        bits.append(f"{pyear}")
    if mode:
        bits.append(mode)
    if ts:
        bits.append(ts)
    return "Supabase · " + " · ".join(bits)


# ---------------------------------------------------------------------------
# Tecscon forecast pipeline API — monthly ledger upload + job polling.
# The operator uploads a month's ledger; we POST it to the pipeline service
# SERVER-SIDE (so the API key never reaches the browser), then poll the job
# until it finishes and reload Supabase. Spec: Tecscon_API_Integration_Guide.
# ---------------------------------------------------------------------------
def _tecscon_creds() -> tuple[str | None, str | None]:
    try:
        url = st.secrets.get("tecscon_api_url")
        key = st.secrets.get("tecscon_api_key")
    except Exception:
        return None, None
    return (str(url).rstrip("/") if url else None, str(key) if key else None)


def _api_upload_ledger(file_bytes: bytes, filename: str, url: str, key: str) -> dict:
    """POST the ledger to /api/process-ledger. Returns the queued-job payload
    ({job_id, year_month, status, message}). Raises on non-2xx / network error."""
    import requests
    resp = requests.post(
        f"{url}/api/process-ledger",
        files={"file": (filename or "ledger.xlsx", file_bytes)},
        headers={"X-API-Key": key},
        timeout=60,
    )
    resp.raise_for_status()
    return resp.json()


def _api_job_status(job_id: str, url: str, key: str) -> dict:
    """GET /api/job-status/{job_id}. Returns the status payload."""
    import requests
    resp = requests.get(
        f"{url}/api/job-status/{job_id}",
        headers={"X-API-Key": key},
        timeout=20,
    )
    resp.raise_for_status()
    return resp.json()


_LEDGER_COLUMNS = ["SN", "ItemCode", "Item Name", "Date", "Type", "Doc No",
                   "Site No and Project", "W/H", "Qty Dlv", "Qty Ret", "Balance"]


def _normalize_ledger(file_bytes: bytes) -> tuple[bytes | None, str | None, dict]:
    """Turn the raw 'customer-wise stock ledger' ERP export into the clean
    single-sheet layout the pipeline expects, so the customer can upload their
    file exactly as exported. We:
      - find the real header row (skipping company/title banner rows),
      - map the 11 required columns by name (ignoring extra cols like Order No,
        LPO No, Ref. No),
      - keep only transaction rows (drop 'Customer' section headers, blank
        separator rows, and dateless subtotal lines),
      - normalise dates to DD/MM/YYYY and renumber SN.
    Returns (clean_xlsx_bytes, error_message, stats)."""
    import io as _io
    import datetime as _dt
    import re as _re
    import openpyxl

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return None, f"could not open the Excel file ({exc.__class__.__name__})", {}
    try:
        ws = wb[wb.sheetnames[0]]
        data = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        return None, f"could not read the sheet ({exc.__class__.__name__})", {}
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # Locate the header row: scan the first 30 rows for one carrying SN + ItemCode.
    header = None
    hdr_i = None
    for i, r in enumerate(data[:30]):
        vals = [str(c).strip() if c is not None else "" for c in r]
        if "SN" in vals and "ItemCode" in vals:
            header, hdr_i = vals, i
            break
    if header is None:
        return None, ("couldn't find the ledger header row (expected columns like SN, "
                      "ItemCode, Item Name, Date...). Is this the stock-ledger export?"), {}

    idx = {n: header.index(n) for n in _LEDGER_COLUMNS if n in header}
    missing = [c for c in _LEDGER_COLUMNS if c not in idx]
    if missing:
        return None, "the ledger is missing expected columns: " + ", ".join(missing), {}

    def fmt_date(v):
        if isinstance(v, (_dt.datetime, _dt.date)):
            return v.strftime("%d/%m/%Y")
        s = str(v).strip()
        m = _re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}" if m else s

    def cell(r, name):
        i = idx[name]
        return r[i] if i < len(r) else None

    out: list[list] = []
    cust = blank = nodate = 0
    for r in data[hdr_i + 1:]:
        item = cell(r, "ItemCode")
        if isinstance(item, str) and item.strip() == "Customer":
            cust += 1
            continue
        if (item is None or str(item).strip() == "") and cell(r, "SN") is None:
            blank += 1
            continue
        d = cell(r, "Date")
        if d is None or str(d).strip() == "":
            nodate += 1
            continue
        out.append([
            len(out) + 1, cell(r, "ItemCode"), cell(r, "Item Name"), fmt_date(d),
            cell(r, "Type"), cell(r, "Doc No"), cell(r, "Site No and Project"),
            cell(r, "W/H"), cell(r, "Qty Dlv"), cell(r, "Qty Ret"), cell(r, "Balance"),
        ])
    if not out:
        return None, "no transaction rows found after cleaning the ledger.", {}

    nwb = openpyxl.Workbook()
    nws = nwb.active
    nws.title = "Sheet1"
    nws.append(_LEDGER_COLUMNS)
    for row in out:
        nws.append(row)
    for rr in range(2, 2 + len(out)):
        nws.cell(row=rr, column=4).number_format = "@"  # keep Date as text
    buf = _io.BytesIO()
    nwb.save(buf)
    months = sorted({row[3][6:10] + "-" + row[3][3:5] for row in out
                     if _re.match(r"^\d{2}/\d{2}/\d{4}$", row[3])})
    return buf.getvalue(), None, {
        "rows": len(out), "skipped_customer": cust, "skipped_blank": blank,
        "skipped_nodate": nodate, "months": months,
    }


def _reload_all_from_supabase() -> bool:
    """Re-pull every prediction from Supabase (busting the short TTL cache) so
    the dashboard reflects rows the pipeline just wrote. True on success."""
    url, key = _supabase_creds()
    if not (url and key):
        return False
    try:
        _fetch_supabase_runs.clear()
        _fetch_all_supabase_predictions.clear()
    except Exception:
        pass
    try:
        runs = _fetch_supabase_runs(url, key)
        records, mape_summary, meta = _load_supabase_all(url, key, runs)
    except Exception:
        return False
    if not records:
        return False
    st.session_state.records = records
    st.session_state.mape_summary = mape_summary
    st.session_state.source_label = _supabase_all_source_label(meta)
    st.session_state.runs_list = runs
    st.session_state.loaded_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    return True


# Component JSX files (top-level helpers + per-screen widgets) that get
# inlined into the served HTML.  Order matters: `data` exposes the
# Supabase hook + mappers used by everything else; `app` is the entry
# component and must be last.
_COMPONENT_FILES = [
    "sidebar", "charts", "explorer", "insights", "upload", "datasource", "actionflow", "newitems", "iteminsight",
]
_TOP_LEVEL_JSX = ["tweaks-panel", "data", "app"]


@st.cache_data(show_spinner=False)
def _read_design_files(version: float) -> dict[str, str]:
    files: dict[str, str] = {}
    for name in _COMPONENT_FILES:
        files[name] = (FRONTEND / "components" / f"{name}.jsx").read_text(encoding="utf-8")
    for name in _TOP_LEVEL_JSX:
        files[name] = (FRONTEND / f"{name}.jsx").read_text(encoding="utf-8")
    files["main_html"] = (FRONTEND / "index.html").read_text(encoding="utf-8")
    return files


def _design_version() -> float:
    paths = [FRONTEND / "index.html"]
    paths += [FRONTEND / f"{n}.jsx" for n in _TOP_LEVEL_JSX]
    paths += [FRONTEND / "components" / f"{n}.jsx" for n in _COMPONENT_FILES]
    return max(p.stat().st_mtime for p in paths)


def build_html(
    records: list[dict],
    source_label: str,
    last_error: str | None,
    mape_summary: list[dict] | None = None,
    runs: list[dict] | None = None,
    current_run_id: int | None = None,
    years: list[dict] | None = None,
    current_year: int | None = None,
) -> str:
    src = _read_design_files(_design_version())
    html = src["main_html"]

    fetch_old = (
        "fetch('data.json').then(r=>r.json()).then(d=>{ window.__RAW_DATA = d;"
        " window.dispatchEvent(new Event('dataready')); });"
    )
    json_str = json.dumps(records, default=str).replace("</", "<\\/")
    mape_str = json.dumps(mape_summary or [], default=str).replace("</", "<\\/")
    runs_str = json.dumps(runs or [], default=str).replace("</", "<\\/")
    years_str = json.dumps(years or [], default=str).replace("</", "<\\/")

    bridge_js = """
window.__showErrorToast = function(msg) {
  if (!msg) return;
  var t = document.createElement('div');
  t.style.cssText = 'position:fixed;top:14px;left:50%;transform:translateX(-50%);background:#FEE2E2;border:1px solid #DC2626;color:#991B1B;padding:10px 16px;border-radius:10px;font:600 12.5px Outfit,system-ui,sans-serif;box-shadow:0 6px 18px rgba(0,0,0,.08);z-index:9999;max-width:520px;line-height:1.5;display:flex;align-items:center;gap:10px;';
  t.innerHTML = '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="#DC2626" stroke-width="2"><circle cx="12" cy="12" r="10"></circle><line x1="12" y1="8" x2="12" y2="12"></line><line x1="12" y1="16" x2="12.01" y2="16"></line></svg><span>'+ msg.replace(/[<>&]/g, function(c){return {'<':'&lt;','>':'&gt;','&':'&amp;'}[c];}) +'</span><button style="margin-left:auto;background:transparent;border:0;color:#991B1B;cursor:pointer;font-size:16px;line-height:1">×</button>';
  document.body.appendChild(t);
  t.querySelector('button').addEventListener('click', function() { t.remove(); });
  setTimeout(function() { if (t.parentNode) t.remove(); }, 8000);
};
window.addEventListener('dataready', function() {
  if (window.__LAST_ERROR) window.__showErrorToast('Upload rejected · ' + window.__LAST_ERROR);
});
window.__resetToBundled = function() {
  try {
    var doc = window.parent.document;
    var input = doc.querySelector('[data-testid="stFileUploaderDropzoneInput"]')
      || doc.querySelector('[data-testid="stFileUploader"] input[type=\\"file\\"]')
      || doc.querySelector('input[type=\\"file\\"]');
    if (!input) return false;
    var blob = new Blob(['__RESET_TO_BUNDLED__'], { type: 'text/csv' });
    var file = new File([blob], '__RESET_TO_BUNDLED__.csv', { type: 'text/csv' });
    var dt = new DataTransfer();
    dt.items.add(file);
    var setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'files').set;
    setter.call(input, dt.files);
    input.dispatchEvent(new Event('change', { bubbles: true }));
    return true;
  } catch (e) { console.error(e); return false; }
};
window.__expandHostUploader = function() {
  try {
    var doc = window.parent.document;
    var details = doc.querySelector('[data-testid="stExpander"] details');
    var summary = doc.querySelector('[data-testid="stExpander"] summary');
    if (summary && details && !details.open) summary.click();
  } catch (e) {}
};
window.__switchRun = function(runId) {
  // Older single-run switcher. Kept so any older client builds keep
  // working — the new picker submits years, not run ids.
  try {
    var doc = window.parent.document;
    var input = doc.querySelector('[data-testid="stFileUploaderDropzoneInput"]')
      || doc.querySelector('[data-testid="stFileUploader"] input[type=\\"file\\"]')
      || doc.querySelector('input[type=\\"file\\"]');
    if (!input) return false;
    var idStr = String(runId);
    var blob = new Blob([idStr], { type: 'text/csv' });
    var file = new File([blob], '__RUN_SWITCH__' + idStr + '.csv', { type: 'text/csv' });
    var dt = new DataTransfer();
    dt.items.add(file);
    var setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'files').set;
    setter.call(input, dt.files);
    input.dispatchEvent(new Event('change', { bubbles: true }));
    return true;
  } catch (e) { console.error('run switch failed', e); return false; }
};
window.__switchYear = function(year) {
  // Year-level switcher. Submits a sentinel-named file whose body holds
  // the integer year. Python-side handler unions every Supabase run that
  // touches that year into a single continuous view (Backtest + Future
  // months distinguished via the per-row forecast_mode column).
  try {
    var doc = window.parent.document;
    var input = doc.querySelector('[data-testid="stFileUploaderDropzoneInput"]')
      || doc.querySelector('[data-testid="stFileUploader"] input[type=\\"file\\"]')
      || doc.querySelector('input[type=\\"file\\"]');
    if (!input) return false;
    var ys = String(year);
    var blob = new Blob([ys], { type: 'text/csv' });
    var file = new File([blob], '__YEAR_SWITCH__' + ys + '.csv', { type: 'text/csv' });
    var dt = new DataTransfer();
    dt.items.add(file);
    var setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'files').set;
    setter.call(input, dt.files);
    input.dispatchEvent(new Event('change', { bubbles: true }));
    return true;
  } catch (e) { console.error('year switch failed', e); return false; }
};
"""
    # SUPABASE_URL / SUPABASE_KEY are exposed to the iframe so the React
    # side can use supabase-js directly for run switching (no Streamlit
    # rerun, no iframe reload). The anon key is meant for client-side use
    # — RLS is enforced server-side, the key just identifies the project.
    sb_url, sb_key = _supabase_creds()
    upload_job = st.session_state.get("upload_job")
    fetch_new = (
        f"window.__RAW_DATA = {json_str}; "
        f"window.__MAPE_SUMMARY = {mape_str}; "
        f"window.__RUNS_LIST = {runs_str}; "
        f"window.__CURRENT_RUN_ID = {json.dumps(current_run_id)}; "
        f"window.__YEARS_LIST = {years_str}; "
        f"window.__CURRENT_YEAR = {json.dumps(current_year)}; "
        f"window.__SOURCE_LABEL = {json.dumps(source_label)}; "
        f"window.__LAST_ERROR = {json.dumps(last_error)}; "
        f"window.__SUPABASE_URL = {json.dumps(sb_url)}; "
        f"window.__SUPABASE_KEY = {json.dumps(sb_key)}; "
        f"window.__UPLOAD_JOB = {json.dumps(upload_job)}; "
        f"{bridge_js} "
        "setTimeout(() => window.dispatchEvent(new Event('dataready')), 0);"
    )
    html = html.replace(fetch_old, fetch_new)

    # Inline every <script type="text/babel" src="..."></script> so the
    # served HTML is fully self-contained (no separate JSX network round
    # trips). The key on the right matches the dictionary built by
    # _read_design_files; the path on the left matches the src attribute
    # in index.html.
    refs = [
        ("tweaks-panel.jsx", "tweaks-panel"),
        ("data.jsx", "data"),
        ("components/sidebar.jsx", "sidebar"),
        ("components/charts.jsx", "charts"),
        ("components/datasource.jsx", "datasource"),
        ("components/explorer.jsx", "explorer"),
        ("components/upload.jsx", "upload"),
        ("components/insights.jsx", "insights"),
        ("components/actionflow.jsx", "actionflow"),
        ("components/newitems.jsx", "newitems"),
        ("components/iteminsight.jsx", "iteminsight"),
        ("app.jsx", "app"),
    ]
    for src_path, key in refs:
        old = f'<script type="text/babel" src="{src_path}"></script>'
        body = src[key].replace("</script>", "<\\/script>")
        new = (
            f'<script type="text/babel" data-src="{src_path}">\n{body}\n</script>'
        )
        html = html.replace(old, new)

    return html


# ---------------------------------------------------------------------------
# Session state defaults — prefer Supabase, fall back to bundled Excel.
# ---------------------------------------------------------------------------
def _bootstrap_session_state() -> None:
    """Populate records / mape_summary / source_label / years on first load.
    Loads the latest predict_year from Supabase by default (merging Backtest
    + Future runs for that year). On any failure falls back to the bundled
    Excel and leaves a non-fatal note in last_error."""
    url, key = _supabase_creds()
    if url and key:
        try:
            runs = _fetch_supabase_runs(url, key)
            if runs:
                records, mape_summary, meta = _load_supabase_all(url, key, runs)
                if records:
                    st.session_state.records = records
                    st.session_state.mape_summary = mape_summary
                    st.session_state.source_label = _supabase_all_source_label(meta)
                    # Years/runs lists are no longer surfaced in the UI — the
                    # year picker is gone — but we keep an empty placeholder so
                    # downstream code doesn't trip on a missing key.
                    st.session_state.years_list = []
                    st.session_state.current_year = None
                    st.session_state.runs_list = runs
                    st.session_state.current_run_id = None
                    st.session_state.loaded_at = datetime.now().strftime("%Y-%m-%d %H:%M")
                    return
            # creds present but no runs / no predictions — fall through to Excel
            st.session_state.last_error = "Supabase returned 0 runs — using bundled data."
        except Exception as exc:
            # Don't block the app: degrade to the bundled Excel and surface the
            # error so the user knows the live DB wasn't reachable.
            st.session_state.last_error = f"Supabase unavailable ({exc.__class__.__name__}); using bundled data."

    # Excel fallback — only if the bundled file exists. Once the app is
    # running fully off Supabase we drop the xlsx from the repo, so this
    # branch is just a safety net for local dev.
    if DEFAULT_DATA.exists():
        all_data = load_default_records(DEFAULT_DATA.stat().st_mtime)
        st.session_state.records = all_data["monthly"]
        st.session_state.mape_summary = all_data["mape"]
        st.session_state.source_label = f"Bundled · {DEFAULT_DATA.name}"
    else:
        # Nothing to fall back to — render an empty dashboard and explain why
        # in the toast so the user knows to configure Supabase creds.
        st.session_state.records = []
        st.session_state.mape_summary = []
        st.session_state.source_label = "No data — configure Supabase in Settings → Secrets"
        if not st.session_state.get("last_error"):
            st.session_state.last_error = (
                "Supabase credentials missing and no bundled Excel found — set "
                "supabase_url + supabase_key in .streamlit/secrets.toml (or "
                "Streamlit Cloud Settings → Secrets)."
            )
    st.session_state.runs_list = []
    st.session_state.current_run_id = None
    st.session_state.loaded_at = datetime.now().strftime("%Y-%m-%d %H:%M")


# Bootstrap on first run, and retry if Supabase is configured but the
# previous attempt ended up empty (transient failure, or a code-edit
# happened mid-bootstrap during local dev). Don't retry once the user
# has uploaded an Excel — that's their explicit override.
def _should_retry_bootstrap() -> bool:
    if "records" not in st.session_state:
        return True
    if st.session_state.records:
        return False
    if (st.session_state.get("source_label") or "").startswith("Uploaded"):
        return False
    return all(_supabase_creds())


if _should_retry_bootstrap():
    # Wipe any stale empty state so the retry uses fresh data.
    for k in ("records", "mape_summary", "source_label", "years_list",
              "current_year", "runs_list", "current_run_id"):
        st.session_state.pop(k, None)
    _bootstrap_session_state()

if "last_error" not in st.session_state:
    st.session_state.last_error = None

# ---------------------------------------------------------------------------
# Bridge file_uploader — must always exist so the login iframe (and the
# dashboard, once authed) can submit through it via the DataTransfer trick.
# ---------------------------------------------------------------------------
with st.container(key="forecast_bridge"):
    upload = st.file_uploader(
        "bridge_uploader",
        type=["xlsx", "csv"],
        label_visibility="collapsed",
        key="data_uploader",
    )

# ---------------------------------------------------------------------------
# Handle login attempts BEFORE anything else. Password arrives as the
# content of a sentinel-named file, never in the filename. We DON'T call
# st.rerun() here — the file_uploader's value persists across reruns until
# a new file is uploaded, so a rerun would re-process the same attempt
# forever. Instead, we set state and let the script flow naturally: the
# gate below either skips (authed) or renders the login iframe with the
# new error.
# ---------------------------------------------------------------------------
if (
    upload is not None
    and upload.name == "__LOGIN_ATTEMPT__.csv"
    and not st.session_state.get("authed")
):
    attempted = upload.getvalue().decode("utf-8", errors="ignore").strip()
    expected = _expected_password()
    if expected is None:
        st.session_state["_login_err"] = (
            "No password is configured on the server. Add `password = \"…\"` to "
            ".streamlit/secrets.toml (or the Streamlit Cloud Secrets section)."
        )
    elif attempted and attempted == expected:
        st.session_state["authed"] = True
        st.session_state.pop("_login_err", None)
    else:
        st.session_state["_login_err"] = "Incorrect password. Try again."

# ---------------------------------------------------------------------------
# Gate: if not authed, render the JSX login iframe and stop.
# (The bridge file_uploader is already mounted above, so the iframe can
# submit through it.)
# ---------------------------------------------------------------------------
if not st.session_state.get("authed"):
    err = st.session_state.get("_login_err")
    _render_login_screen(error=err)
    st.stop()

# ---------------------------------------------------------------------------
# Authed path: dashboard styles + data upload handling + dashboard render.
# ---------------------------------------------------------------------------
st.markdown(_DASHBOARD_CSS, unsafe_allow_html=True)

if upload is not None:
    if upload.name == "__RESET_TO_BUNDLED__.csv":
        # "Reset" now means "reload everything from Supabase".
        url, key = _supabase_creds()
        if url and key:
            try:
                runs = _fetch_supabase_runs(url, key)
                if runs:
                    records, mape_summary, meta = _load_supabase_all(url, key, runs)
                    st.session_state.records = records
                    st.session_state.mape_summary = mape_summary
                    st.session_state.source_label = _supabase_all_source_label(meta)
                    st.session_state.years_list = []
                    st.session_state.current_year = None
                    st.session_state.runs_list = runs
                    st.session_state.current_run_id = None
                    st.session_state.loaded_at = datetime.now().strftime("%Y-%m-%d %H:%M")
                    st.session_state.last_error = None
                else:
                    st.session_state.last_error = "Supabase has no runs to reset to."
            except Exception as exc:
                st.session_state.last_error = f"Could not reset from Supabase: {exc}"
        elif DEFAULT_DATA.exists():
            all_data = load_default_records(DEFAULT_DATA.stat().st_mtime)
            st.session_state.records = all_data["monthly"]
            st.session_state.mape_summary = all_data["mape"]
            st.session_state.source_label = f"Bundled · {DEFAULT_DATA.name}"
            st.session_state.years_list = []
            st.session_state.current_year = None
            st.session_state.runs_list = []
            st.session_state.current_run_id = None
            st.session_state.loaded_at = datetime.now().strftime("%Y-%m-%d %H:%M")
            st.session_state.last_error = None
        else:
            st.session_state.last_error = "Nothing to reset to — Supabase not configured and no bundled Excel."
    elif upload.name.startswith("__YEAR_SWITCH__") and upload.name.endswith(".csv"):
        # Switch to a different forecast year — fetches all rows where
        # year_month falls in the requested year, merging Backtest +
        # Future runs for that year into a single continuous view.
        try:
            new_year = int(upload.getvalue().decode("utf-8", errors="ignore").strip())
        except ValueError:
            new_year = None
        url, key = _supabase_creds()
        if new_year is None or not (url and key):
            st.session_state.last_error = "Could not switch year — bad year value or missing Supabase credentials."
        else:
            try:
                runs = _fetch_supabase_runs(url, key)
                records, mape_summary, year_meta = _load_supabase_year(url, key, new_year, runs)
                if not records:
                    st.session_state.last_error = f"No predictions found for {new_year}."
                else:
                    pred_cache = {new_year: _fetch_supabase_predictions_by_year(url, key, new_year)}
                    st.session_state.records = records
                    st.session_state.mape_summary = mape_summary
                    st.session_state.source_label = _supabase_year_source_label(year_meta)
                    st.session_state.years_list = _years_summary(runs, pred_cache)
                    st.session_state.current_year = new_year
                    st.session_state.loaded_at = datetime.now().strftime("%Y-%m-%d %H:%M")
                    st.session_state.last_error = None
            except Exception as exc:
                st.session_state.last_error = f"Failed to load year {new_year}: {exc}"
    elif upload.name.startswith("__RUN_SWITCH__") and upload.name.endswith(".csv"):
        # Older single-run switch sentinel — keep working for back-compat
        # but no UI surfaces it any more.
        try:
            new_run_id = int(upload.getvalue().decode("utf-8", errors="ignore").strip())
        except ValueError:
            new_run_id = None
        url, key = _supabase_creds()
        if new_run_id is None or not (url and key):
            st.session_state.last_error = "Could not switch run — bad run id or missing Supabase credentials."
        else:
            try:
                records, mape_summary, run_meta = _load_supabase_run(url, key, new_run_id)
                if not records:
                    st.session_state.last_error = f"Run #{new_run_id} has no predictions."
                else:
                    st.session_state.records = records
                    st.session_state.mape_summary = mape_summary
                    st.session_state.source_label = _supabase_source_label(run_meta)
                    st.session_state.current_run_id = new_run_id
                    st.session_state.loaded_at = datetime.now().strftime("%Y-%m-%d %H:%M")
                    st.session_state.last_error = None
            except Exception as exc:
                st.session_state.last_error = f"Failed to load run #{new_run_id}: {exc}"
    elif upload.name.startswith("__LEDGER_UPLOAD__"):
        # Operator uploaded a monthly ledger. Hand it to the forecast pipeline
        # API (server-side, key stays in secrets) and start tracking the job.
        # Guard on file_id so a persisted upload isn't re-POSTed on later reruns.
        fid = getattr(upload, "file_id", None)
        if st.session_state.get("_ledger_fid") != fid:
            st.session_state["_ledger_fid"] = fid
            api_url, api_key = _tecscon_creds()
            if not (api_url and api_key):
                # Report exactly which secrets THIS running instance can see (key
                # names only, never values) so a "but I added it!" can be settled
                # instantly — if tecscon_* aren't listed, they aren't on this app.
                try:
                    seen = [k for k in ("password", "supabase_url", "supabase_key",
                                        "tecscon_api_url", "tecscon_api_key") if k in st.secrets]
                except Exception:
                    seen = []
                st.session_state.last_error = (
                    "Upload pipeline isn't configured on this deployment. Secrets this app "
                    "instance can currently see: " + (", ".join(seen) if seen else "none")
                    + ". If tecscon_api_url / tecscon_api_key aren't in that list, add them to "
                    "THIS app's Settings -> Secrets, click Save, and reboot the app."
                )
            else:
                # Send the raw ledger straight to the pipeline, untouched. The
                # Azure side parses and validates the customer's export format,
                # so we do NOT transform or schema-check it here.
                orig = upload.name[len("__LEDGER_UPLOAD__"):].strip() or "ledger.xlsx"
                try:
                    res = _api_upload_ledger(upload.getvalue(), orig, api_url, api_key)
                    st.session_state.upload_job = {
                        "job_id": res.get("job_id"),
                        "year_month": res.get("year_month"),
                        "status": res.get("status", "queued"),
                        "current_step": res.get("message"),
                        "error_message": None,
                        "file_name": orig,
                        "started_at": datetime.now().strftime("%H:%M"),
                    }
                    st.session_state.last_error = None
                except Exception as exc:
                    st.session_state.last_error = f"Could not start the forecast pipeline: {exc}"
    elif upload.name.startswith("__JOB_CLEAR__"):
        st.session_state.pop("upload_job", None)
    # NOTE: the old "upload a results Excel and preview it" path (which schema-
    # checked the file for a Monthly_Predictions sheet) has been removed. The
    # only upload now is the monthly ledger, which is forwarded raw to the
    # pipeline (handled by the __LEDGER_UPLOAD__ branch above); validation of
    # the ledger format happens on the Azure side.


last_error = st.session_state.last_error
st.session_state.last_error = None
html_doc = build_html(
    st.session_state.records,
    st.session_state.source_label,
    last_error,
    mape_summary=st.session_state.get("mape_summary", []),
    runs=st.session_state.get("runs_list", []),
    current_run_id=st.session_state.get("current_run_id"),
    years=st.session_state.get("years_list", []),
    current_year=st.session_state.get("current_year"),
)
components_html(html_doc, height=1100, scrolling=True)

# ---------------------------------------------------------------------------
# Poll an in-flight pipeline job. The fragment reruns on its own every 30s
# WITHOUT re-rendering the dashboard iframe (so the 12-18 min wait isn't a
# flicker-fest); only when the job finishes does it reload Supabase and do a
# single full rerun so the fresh predictions appear.
# ---------------------------------------------------------------------------
_job = st.session_state.get("upload_job")
if _job and _job.get("status") in ("queued", "running") and all(_tecscon_creds()):

    @st.fragment(run_every="30s")
    def _poll_upload_job():
        j = st.session_state.get("upload_job")
        if not j or j.get("status") not in ("queued", "running"):
            return
        a_url, a_key = _tecscon_creds()
        try:
            s = _api_job_status(j["job_id"], a_url, a_key)
        except Exception:
            return  # transient network error — try again on the next tick
        if not s:
            return
        j["status"] = s.get("status", j["status"])
        j["current_step"] = s.get("current_step")
        j["error_message"] = s.get("error_message")
        j["year_month"] = s.get("year_month", j.get("year_month"))
        st.session_state.upload_job = j
        if j["status"] == "complete":
            _reload_all_from_supabase()
            st.rerun()
        elif j["status"] == "failed":
            st.session_state.last_error = "Forecast pipeline failed: " + (j.get("error_message") or "unknown error")
            st.rerun()

    _poll_upload_job()
